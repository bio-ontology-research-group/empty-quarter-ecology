#!/usr/bin/env python3
"""Build the extended extraction-batch map (Trip 4 + Trip 5 blanks).

Reads Marwa Abdelhakim's Trip 4 EB-sample map (mail of 2026-08-26) and the
Trip 5 workbook already used by the published screen, resolves every listed
sample ID against the column names of the canonical feature table, and writes:

  inputs/extraction_batch_map_extended.tsv   one row per (blank, sample)
  inputs/blank_libraries.tsv                 one row per blank with its status
  inputs/mapping_report.json                 mismatches and counts

The original metadata files are not modified.
"""
from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
ROOT = HERE.parents[1]
TRIP4_XLSX = HERE / "inputs" / "EB_Sample_Map_FourthTrip2 correct.xlsx"
TRIP5_XLSX = ROOT / "data/metadata/samples/Sequenced_Samples_by_EB_FifthTrip.xlsx"
CANONICAL = ROOT / "data/processed/taxonomy/taxon-tables/feature-table-trips1-5.tsv"
ADDITIONAL = ROOT / "data/metadata/samplesheets/additional_fastqs_v2.tsv"
TRIP5_SHEET = (
    ROOT.parent / "data-paper/evidence/controls/source_snapshots/ibex_trip5_16s_samplesheet.tsv"
)


def split_ids(value):
    if value is None:
        return []
    return [p.strip() for p in str(value).split(",") if p.strip()]


def canonical_columns():
    with CANONICAL.open(encoding="utf-8") as fh:
        fh.readline()
        return fh.readline().rstrip("\n").split("\t")[1:]


def body(col):
    m = re.match(r"^e\d+_(.+)$", col)
    return m.group(1) if m else col


def main():
    columns = canonical_columns()
    by_body = defaultdict(list)
    for c in columns:
        by_body[body(c)].append(c)

    # Trip 4 map from Marwa's workbook
    ws = load_workbook(TRIP4_XLSX, read_only=True, data_only=True).active
    trip4 = {}
    unsequenced_note = ""
    for row in ws.iter_rows(values_only=True):
        label = str(row[0]).strip() if row[0] else ""
        if re.fullmatch(r"EB\d*", label):
            trip4[label] = {
                "index_code": str(row[1]).strip(),
                "declared_count": int(row[2]) if row[2] is not None else None,
                "sites_note": str(row[3]) if row[3] else "",
                "samples": split_ids(row[4]),
            }
        elif label.startswith("PCR"):
            trip4_pcr_index = str(row[1]).strip()
        elif "wasn't sequenced" in label:
            unsequenced_note = label
    # Trip 5 map (as parsed by the published screen)
    ws5 = load_workbook(TRIP5_XLSX, read_only=True, data_only=True).active
    trip5 = {}
    for row in ws5.iter_rows(values_only=True):
        label = str(row[0]).strip() if row[0] else ""
        if label.startswith("EB") and label[2:].isdigit() and 1 <= int(label[2:]) <= 17:
            trip5[label] = {
                "date": datetime.strptime(str(row[1]), "%m/%d/%Y").date().isoformat(),
                "index_code": str(row[2]).strip().replace("  ", " "),
                "declared_count": int(row[3]),
                "samples": split_ids(row[4]),
            }

    # Where do the Trip 4 blank libraries live?
    additional = {}
    with ADDITIONAL.open(encoding="utf-8") as fh:
        for r in csv.DictReader(fh, delimiter="\t"):
            additional[r["Sample Name"]] = r["FASTQ Paths"].split(";")[0]
    trip4_library_names = {"EB": "SEB"}  # unnumbered Trip 4 EB is library 'SEB' (index SU0168)
    for k in trip4:
        trip4_library_names.setdefault(k, k)

    rows, blanks, report = [], [], {"trip4": {}, "trip5": {}}
    dup = Counter()
    for trip, mapping, dflt_col in (("Trip4", trip4, None), ("Trip5", trip5, None)):
        for blank, rec in mapping.items():
            for s in rec["samples"]:
                dup[(trip, s)] += 1
    for trip, mapping in (("Trip4", trip4), ("Trip5", trip5)):
        unmatched, ambiguous, matched = [], [], 0
        for blank, rec in mapping.items():
            if trip == "Trip4":
                lib = trip4_library_names[blank]
                m = re.search(r"SU(\d{4})", additional.get(lib, ""))
                lib_index = f"xGen 10nt UDI Index Pair {int(m.group(1))}" if m else ""
                # The canonical table's EB1-EB5 columns are the Trip 5 blanks
                # (index pairs 307/410/319/331/422 from the Trip 5 samplesheet);
                # the Trip 4 libraries EB1-EB5 (index pairs 216-264, run
                # novaseq_14_07_25) share the label but were never denoised.
                blank_col = f"TRIP4_{lib}"  # column name used for the rerun table
                status = (
                    "sequenced_never_denoised_fastq_on_ibex" if lib in additional else
                    "no_fastq_record"
                )
                blanks.append({
                    "trip": trip, "blank_label": blank, "library_name": lib,
                    "workbook_index_code": rec["index_code"],
                    "fastq_index_code": lib_index,
                    "index_agrees": str(lib_index == rec["index_code"]),
                    "extraction_date": "",
                    "declared_sample_count": rec["declared_count"],
                    "listed_sample_count": len(rec["samples"]),
                    "table_column": blank_col, "status": status,
                    "fastq_r1": additional.get(lib, ""),
                })
            else:
                in_table = blank in columns
                blank_col = blank
                blanks.append({
                    "trip": trip, "blank_label": blank, "library_name": blank,
                    "workbook_index_code": rec["index_code"], "fastq_index_code": rec["index_code"],
                    "index_agrees": "True", "extraction_date": rec["date"],
                    "declared_sample_count": rec["declared_count"],
                    "listed_sample_count": len(rec["samples"]),
                    "table_column": blank_col,
                    "status": "in_canonical_table" if in_table else "absent_from_canonical_table",
                    "fastq_r1": "",
                })
            for s in rec["samples"]:
                cols = by_body.get(s, [])
                if len(cols) == 1:
                    matched += 1
                    col = cols[0]
                elif not cols:
                    unmatched.append(f"{blank}:{s}")
                    col = ""
                else:
                    ambiguous.append(f"{blank}:{s}:{cols}")
                    col = ""
                rows.append({
                    "trip": trip, "extraction_blank": blank, "blank_table_column": blank_col,
                    "sample_id": s, "canonical_column": col,
                    "in_canonical_table": str(bool(col)),
                    "duplicate_assignment": str(dup[(trip, s)] > 1),
                })
        report[trip.lower()] = {
            "blanks": list(mapping),
            "listed_samples": sum(len(r["samples"]) for r in mapping.values()),
            "matched_in_canonical_table": matched,
            "unmatched": unmatched, "ambiguous": ambiguous,
            "duplicate_assignments": sorted(s for (t, s), n in dup.items() if t == trip and n > 1),
            "declared_vs_listed_mismatch": {
                b: (r["declared_count"], len(r["samples"]))
                for b, r in mapping.items() if r["declared_count"] != len(r["samples"])
            },
        }
    # Trip 4 profiles in the canonical table that no sequenced blank covers
    t4_cols = [c for c in columns if re.fullmatch(r"e\d+_S\d+(PR|S|D)r\d+", c)]
    mapped_bodies = {r["sample_id"] for r in rows if r["trip"] == "Trip4"}
    uncovered = sorted(body(c) for c in t4_cols if body(c) not in mapped_bodies)
    report["trip4"]["canonical_trip4_profiles"] = len(t4_cols)
    report["trip4"]["canonical_trip4_profiles_without_sequenced_blank"] = len(uncovered)
    report["trip4"]["uncovered_profiles"] = uncovered
    report["trip4"]["uncovered_sites"] = sorted({int(re.match(r"S(\d+)", u).group(1)) for u in uncovered})
    report["trip4"]["workbook_unsequenced_note"] = unsequenced_note
    report["trip4"]["pcr_blank_index_code"] = trip4_pcr_index
    report["trip4_blank_profiles_in_canonical_table"] = []
    report["trip4_blank_label_collision"] = (
        "Trip 4 libraries EB1-EB5 share labels with Trip 5 blanks EB1-EB5; the canonical "
        "EB1-EB5 columns are the Trip 5 libraries (ibex_trip5_16s_samplesheet.tsv)"
    )
    report["canonical_control_columns"] = [
        c for c in columns if re.fullmatch(r"(EB\d+|Negative\d+)", c)
    ]

    out = HERE / "inputs"
    with (out / "extraction_batch_map_extended.tsv").open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0]), delimiter="\t", lineterminator="\n")
        w.writeheader(); w.writerows(rows)
    with (out / "blank_libraries.tsv").open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(blanks[0]), delimiter="\t", lineterminator="\n")
        w.writeheader(); w.writerows(blanks)
    (out / "mapping_report.json").write_text(json.dumps(report, indent=2) + "\n")
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
